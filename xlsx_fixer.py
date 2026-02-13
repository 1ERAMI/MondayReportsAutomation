"""
Excel File Fixer Module
Fixes default column width issues in Excel files
"""

from openpyxl import load_workbook


class XLSXFixer:
    """Class to fix Excel file formatting issues"""
    
    @staticmethod
    def fix_default_col_width(file_path):
        """
        Fixes default column width issues in Excel files.
        
        Args:
            file_path (str): Path to the Excel file to fix
        """
        try:
            # Load the workbook
            wb = load_workbook(file_path)
            
            # Iterate through all worksheets
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # Auto-adjust column widths based on content
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            if cell.value:
                                # Calculate the length of the cell value
                                cell_length = len(str(cell.value))
                                if cell_length > max_length:
                                    max_length = cell_length
                        except:
                            pass
                    
                    # Set the column width with some padding
                    # Add 2 for padding, max width of 50 to prevent overly wide columns
                    adjusted_width = min(max_length + 2, 50)
                    if adjusted_width > 0:
                        ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save the workbook
            wb.save(file_path)
            print(f"Fixed column widths in: {file_path}")
            
        except Exception as e:
            print(f"Error fixing column widths: {e}")
            raise


if __name__ == "__main__":
    # Test the fixer
    import sys
    if len(sys.argv) > 1:
        file_path = sys.argv[1]
        print(f"Fixing column widths for: {file_path}")
        XLSXFixer.fix_default_col_width(file_path)
    else:
        print("Usage: python Fix_defaultColWidthPt.py <excel_file_path>")
