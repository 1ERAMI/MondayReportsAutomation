#Camerons Crump & Other Monday Report Automation

#  %%
import RepAutoGmail as RAGA  # Import your Gmail authentication module
from Fix_defaultColWidthPt import XLSXFixer
from DriveUploader import upload_folder_to_drive
import os
import time
import re
import pandas as pd
import base64
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from googleapiclient.http import MediaFileUpload
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
from base64 import urlsafe_b64decode
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
import win32com.client as win32
import pythoncom
import shutil
import tempfile
import traceback
from datetime import datetime, timedelta
import ttkbootstrap as ttk
from ttkbootstrap.constants import *
from tkinter import StringVar, BooleanVar, messagebox
import threading

# import win32com.client
# import win32com.client.makepy

# win32com.client.makepy.main()


# win32.gencache.EnsureDispatch('Excel.Application')
# excel = win32.Dispatch('Excel.Application')
gservice = RAGA.confirm_auth()

# Dynamic path that works on any Windows machine
SAVE_DIRECTORY = os.path.join(os.path.expanduser("~"), "Desktop", "Working", "Python Outputs", "Malissa")


def get_target_date():
    """
    Get today's date if it's Monday, otherwise get the most recent Monday.
    Returns date in YYYY/MM/DD format for Gmail query.
    """
    today = datetime.now().date()
    weekday = today.weekday()  # 0 = Monday, 6 = Sunday

    if weekday == 0:  # Today is Monday
        target_date = today
    else:  # Calculate most recent Monday
        days_since_monday = weekday
        target_date = today - timedelta(days=days_since_monday)

    return target_date.strftime("%Y/%m/%d")


def clear_save_directory(directory):
    """
    Clears all files in the specified directory.
    """
    if os.path.exists(directory):
        shutil.rmtree(directory)
        os.makedirs(directory, exist_ok=True)
        print(f"Cleared directory: {directory}")
    else:
        os.makedirs(directory)
        print(f"Created new directory: {directory}")


# %%
# Helper to sanitize filenames
def sanitize_filename(filename):
    return re.sub(r'[\\/*?:"<>|]', "_", filename)

# Function to retrieve email and download the attachment
def get_report_email(gservice, subject_filter):
    target_date = get_target_date()
    query = f"subject:\"{subject_filter}\" after:{target_date}"
    print(f"Searching for emails after {target_date}")
    results = gservice.users().messages().list(userId="me", q=query).execute()
    messages = results.get("messages", [])

    if not messages:
        print("No emails found with the given subject.")
        return None

    for message in messages:
        msg = gservice.users().messages().get(userId="me", id=message["id"]).execute()
        payload = msg.get("payload", {})
        parts = payload.get("parts", [])

        for part in parts:
            if part.get("filename") and part.get("mimeType") == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
                filename = sanitize_filename(part["filename"])
                body = part.get("body", {})
                if "attachmentId" in body:
                    attachment_id = body["attachmentId"]
                    attachment = gservice.users().messages().attachments().get(
                        userId="me", messageId=message["id"], id=attachment_id
                    ).execute()
                    file_data = base64.urlsafe_b64decode(attachment["data"])

                    # Save the file
                    os.makedirs(SAVE_DIRECTORY, exist_ok=True)
                    file_path = os.path.join(SAVE_DIRECTORY, filename)
                    with open(file_path, "wb") as f:
                        f.write(file_data)
                    print(f"File downloaded: {file_path}")
                    return file_path

    print("No attachments found.")
    return None

# Process Excel file
def process_excel(file_path):
    """
    Fix Excel issues, clean data, and set date columns to Excel 'Short Date' format.
    """
    try:
        print("Fixing Excel structure...")
        XLSXFixer.fix_default_col_width(file_path)

        print("Processing Excel data...")
        df = pd.read_excel(file_path)

        # Define date columns
        date_columns = ["E-Sign Signed Date", "Lead Created Date", "Date of Birth"]
        
        # Convert columns to datetime
        for col in date_columns:
            if col in df.columns:
                df[col] = pd.to_datetime(df[col], errors="coerce")

        # Save DataFrame back to Excel
        df.to_excel(file_path, index=False, engine="openpyxl")

        # Now set Excel's native 'Short Date' format (mm/dd/yyyy) using openpyxl
        wb = load_workbook(file_path)
        ws = wb.active
        
        for col_name in date_columns:
            if col_name in df.columns:
                col_idx = df.columns.get_loc(col_name) + 1  # pandas is 0-based, Excel 1-based
                for cell in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                    if cell[0].value:
                        cell[0].number_format = 'MM/DD/YYYY'  # Excel short date format

        wb.save(file_path)
        print(f"Dates formatted as 'Short Date'. Final file saved: {file_path}")

        return file_path

    except Exception as e:
        print(f"Error processing the Excel file: {e}")
        return None

# Add Excel table to the first sheet
def add_table_to_sheet(file_path, sheet_name):
    try:
        wb = load_workbook(file_path)
        ws = wb[sheet_name]
        max_row, max_col = ws.max_row, ws.max_column
        table_range = f"A1:{ws.cell(row=max_row, column=max_col).coordinate}"
        table = Table(displayName="Table1", ref=table_range)
        style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True, showColumnStripes=True)
        table.tableStyleInfo = style
        ws.add_table(table)
        wb.save(file_path)
        print(f"Table added to sheet: {sheet_name}")

    except Exception as e:
        print(f"Error adding table to sheet: {e}")

def rename_sheet(file_path, old_name, new_name):
    try:
        wb = load_workbook(file_path)
        if old_name in wb.sheetnames:
            wb[old_name].title = new_name
            wb.save(file_path)
            print(f"Sheet '{old_name}' renamed to '{new_name}'.")
        else:
            print(f"Sheet '{old_name}' not found in the workbook.")
    except Exception as e:
        print(f"Error renaming sheet: {e}")

# Create Excel pivot tables
def create_multiple_pivot_tables(file_path, data_sheet_name, pivot_sheet_names):
    excel = None
    try:
        # Initialize COM for this thread
        pythoncom.CoInitialize()

        # clear_gen_py_cache()
        excel = win32.gencache.EnsureDispatch("Excel.Application")
        excel.Visible = True  # Show Excel for debugging

        print(f"Opening workbook: {file_path}")
        wb = excel.Workbooks.Open(os.path.abspath(file_path))
        data_sheet = wb.Worksheets(data_sheet_name)
        print(f"Opened sheet: {data_sheet_name}")

        last_row = data_sheet.UsedRange.Rows.Count
        last_col = data_sheet.UsedRange.Columns.Count
        print(f"Detected used range: {last_row} rows, {last_col} columns")

        headers = [data_sheet.Cells(1, col).Value for col in range(1, last_col + 1)]
        print("Headers found in row 1:", headers)

        if "Status" not in headers:
            raise ValueError("The 'Status' column is missing from the data. Cannot create pivot table.")

        data_range = data_sheet.Range(data_sheet.Cells(1, 1), data_sheet.Cells(last_row, last_col))
        pivot_cache = wb.PivotCaches().Create(SourceType=1, SourceData=data_range)

        for pivot_sheet_name in pivot_sheet_names:
            pivot_sheet = wb.Sheets.Add()
            pivot_sheet.Name = pivot_sheet_name
            print(f"Creating pivot table in new sheet: {pivot_sheet_name}")

            pivot_table = pivot_cache.CreatePivotTable(
                TableDestination=pivot_sheet.Cells(1, 1),
                TableName=f"PivotTable_{pivot_sheet_name.replace(' ', '_')}",
            )

            print("Adding fields to pivot table...")

            status_field_row = pivot_table.PivotFields("Status")
            status_field_row.Orientation = 1  # xlRowField
            status_field_row.Position = 1

            status_field_value = pivot_table.PivotFields("Status")
            status_field_value.Orientation = 4  # xlDataField
            status_field_value.Function = -4112  # xlCount
            status_field_value.Name = "Count of Status"

            print(f"Pivot table '{pivot_sheet_name}' created successfully.")

        wb.Worksheets(data_sheet_name).Activate()
        excel.ActiveWindow.View = 1
        excel.ActiveWindow.SelectedSheets(1).Select()

        wb.Save()
        print("All pivot tables created and workbook saved.")

    except Exception as e:
        print(f"Error creating pivot tables: {e}")

    finally:
        if excel is not None:
            excel.Quit()
        # Uninitialize COM
        pythoncom.CoUninitialize()

# Format and reorder sheets
def format_and_reorder_sheets(file_path, sheet_order):
    """
    Adjust column widths for the first sheet in the order and reorder sheets.
    Ensures only the first sheet is selected upon saving.
    :param file_path: Path to the Excel file.
    :param sheet_order: List of sheet names in the desired order.
    """
    excel = None
    try:
        # Initialize COM for this thread
        pythoncom.CoInitialize()

        excel = win32.gencache.EnsureDispatch("Excel.Application")
        excel.Visible = False
        print(f"Opening workbook for formatting and reordering: {file_path}")

        wb = excel.Workbooks.Open(os.path.abspath(file_path))

        # Verify sheets exist
        existing_sheets = [sheet.Name for sheet in wb.Sheets]
        for sheet in sheet_order:
            if sheet not in existing_sheets:
                raise ValueError(f"Sheet '{sheet}' not found in workbook.")

        # Adjust column widths for the first sheet
        first_sheet_name = sheet_order[0]
        first_sheet = wb.Worksheets(first_sheet_name)
        first_sheet.Activate()

        print(f"Auto-adjusting column widths for '{first_sheet_name}'...")
        first_sheet.UsedRange.Columns.AutoFit()

        # Reorder sheets explicitly
        print("Reordering sheets...")
        for idx, sheet_name in enumerate(sheet_order, start=1):
            wb.Worksheets(sheet_name).Move(Before=wb.Worksheets(idx))

        # Explicitly select only the first sheet and ungroup sheets
        first_sheet.Select()
        excel.ActiveWindow.View = 1  # xlNormalView ensures ungrouped sheets
        excel.ActiveWindow.SelectedSheets(1).Select()

        wb.Save()
        wb.Close()
        print(f"Sheets reordered and column widths adjusted successfully in '{file_path}'.")

    except PermissionError as e:
        print(f"Permission denied error: {e}")
        print("Ensure the Excel file isn't open elsewhere.")
    except Exception as e:
        print(f"Error during formatting and reordering: {e}")
    finally:
        if excel is not None:
            excel.Quit()
        # Uninitialize COM
        pythoncom.CoUninitialize()


def send_email_with_attachments(gservice, to_emails, subject, body, attachment_folder):
    try:
        if isinstance(to_emails, list):
            to_emails = ", ".join(to_emails)

        message = MIMEMultipart()
        message["to"] = to_emails
        message["subject"] = subject
        message.attach(MIMEText(body, "plain"))

        # Attach all .xlsx files in the directory
        for file_name in os.listdir(attachment_folder):
            if file_name.endswith(".xlsx"):
                file_path = os.path.join(attachment_folder, file_name)
                with open(file_path, "rb") as attachment:
                    part = MIMEBase("application", "octet-stream")
                    part.set_payload(attachment.read())
                encoders.encode_base64(part)
                part.add_header(
                    "Content-Disposition",
                    f"attachment; filename={os.path.basename(file_path)}",
                )
                message.attach(part)

        raw_message = base64.urlsafe_b64encode(message.as_bytes()).decode()
        gservice.users().messages().send(userId="me", body={"raw": raw_message}).execute()
        print("Email sent successfully!")
    except Exception as e:
        print(f"Error sending email: {e}")

# def clear_gen_py_cache():
#     """
#     Deletes the 'gen_py' folder in the user's local temp directory to avoid
#     known win32com CLSIDToClassMap errors.
#     """
#     gen_py_path = os.path.join(tempfile.gettempdir(), "gen_py")
#     if os.path.exists(gen_py_path):
#         try:
#             shutil.rmtree(gen_py_path)
#             print(f"Deleted cached COM folder: {gen_py_path}")
#         except Exception as e:
#             print(f"Failed to delete gen_py: {e}")


class EmailSenderUI:
    """Modern UI for selecting email recipients and sending reports"""
    
    def __init__(self, available_emails, subject_filters):
        self.available_emails = available_emails
        self.subject_filters = subject_filters
        self.selected_emails = []
        self.email_vars = {}
        
        # Create main window
        self.root = ttk.Window(themename="darkly")
        self.root.title("Monday Reports - Email Sender")
        self.root.geometry("600x700")
        
        # Configure grid weights
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(1, weight=1)
        
        self.setup_ui()
        
    def setup_ui(self):
        """Setup the UI components"""
        
        # Header
        header_frame = ttk.Frame(self.root, padding=20)
        header_frame.grid(row=0, column=0, sticky="ew", padx=10, pady=10)
        
        title_label = ttk.Label(
            header_frame,
            text="📧 Malissa Reports Sender",
            font=("Segoe UI", 18, "bold"),
            bootstyle="inverse-primary"
        )
        title_label.pack()
        
        subtitle_label = ttk.Label(
            header_frame,
            text="Select recipients and send processed reports",
            font=("Segoe UI", 10)
        )
        subtitle_label.pack(pady=(5, 0))
        
        # Main content area with scrollbar
        content_frame = ttk.Frame(self.root)
        content_frame.grid(row=1, column=0, sticky="nsew", padx=10, pady=10)
        content_frame.columnconfigure(0, weight=1)
        content_frame.rowconfigure(0, weight=1)
        
        # Canvas and scrollbar for email list
        canvas = ttk.Canvas(content_frame)
        scrollbar = ttk.Scrollbar(content_frame, orient="vertical", command=canvas.yview, bootstyle="primary-round")
        scrollable_frame = ttk.Frame(canvas)
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # Email selection section
        email_label = ttk.Label(
            scrollable_frame,
            text="📬 Available Recipients:",
            font=("Segoe UI", 12, "bold"),
            bootstyle="info"
        )
        email_label.pack(anchor="w", padx=20, pady=(10, 5))
        
        # Create checkboxes for each email
        for email in self.available_emails:
            var = BooleanVar(value=False)
            self.email_vars[email] = var
            
            cb_frame = ttk.Frame(scrollable_frame)
            cb_frame.pack(fill="x", padx=30, pady=2)
            
            cb = ttk.Checkbutton(
                cb_frame,
                text=email,
                variable=var,
                bootstyle="primary-round-toggle"
            )
            cb.pack(anchor="w")
        
        # Separator
        sep = ttk.Separator(scrollable_frame, bootstyle="secondary")
        sep.pack(fill="x", padx=20, pady=15)
        
        # Reports section
        reports_label = ttk.Label(
            scrollable_frame,
            text="📊 Reports to Process:",
            font=("Segoe UI", 12, "bold"),
            bootstyle="info"
        )
        reports_label.pack(anchor="w", padx=20, pady=(5, 5))
        
        # Show count of reports
        count_label = ttk.Label(
            scrollable_frame,
            text=f"{len(self.subject_filters)} reports configured",
            font=("Segoe UI", 9)
        )
        count_label.pack(anchor="w", padx=30, pady=(0, 10))
        
        # Pack canvas and scrollbar
        canvas.grid(row=0, column=0, sticky="nsew")
        scrollbar.grid(row=0, column=1, sticky="ns")
        
        # Status bar
        self.status_var = StringVar(value="Ready to send reports")
        self.status_label = ttk.Label(
            self.root,
            textvariable=self.status_var,
            font=("Segoe UI", 9),
            bootstyle="inverse-secondary",
            padding=10
        )
        self.status_label.grid(row=2, column=0, sticky="ew", padx=10, pady=(0, 5))
        
        # Buttons frame
        button_frame = ttk.Frame(self.root, padding=10)
        button_frame.grid(row=3, column=0, sticky="ew", padx=10, pady=10)
        
        # Select/Deselect All buttons
        select_all_btn = ttk.Button(
            button_frame,
            text="✓ Select All",
            command=self.select_all,
            bootstyle="info-outline",
            width=15
        )
        select_all_btn.pack(side="left", padx=5)
        
        deselect_all_btn = ttk.Button(
            button_frame,
            text="✗ Deselect All",
            command=self.deselect_all,
            bootstyle="secondary-outline",
            width=15
        )
        deselect_all_btn.pack(side="left", padx=5)
        
        # Send button
        self.send_btn = ttk.Button(
            button_frame,
            text="🚀 Send Reports",
            command=self.on_send,
            bootstyle="success",
            width=20
        )
        self.send_btn.pack(side="right", padx=5)
        
        # Cancel button
        cancel_btn = ttk.Button(
            button_frame,
            text="Cancel",
            command=self.root.quit,
            bootstyle="danger-outline",
            width=15
        )
        cancel_btn.pack(side="right", padx=5)
    
    def select_all(self):
        """Select all email checkboxes"""
        for var in self.email_vars.values():
            var.set(True)
        self.status_var.set(f"Selected all {len(self.email_vars)} recipients")
    
    def deselect_all(self):
        """Deselect all email checkboxes"""
        for var in self.email_vars.values():
            var.set(False)
        self.status_var.set("All recipients deselected")
    
    def on_send(self):
        """Handle send button click"""
        # Get selected emails
        self.selected_emails = [email for email, var in self.email_vars.items() if var.get()]
        
        if not self.selected_emails:
            messagebox.showwarning(
                "No Recipients",
                "Please select at least one recipient."
            )
            return
        
        # Confirm send
        count = len(self.selected_emails)
        confirm = messagebox.askyesno(
            "Confirm Send",
            f"Send reports to {count} recipient(s)?\n\n" + "\n".join(self.selected_emails)
        )
        
        if confirm:
            self.send_btn.config(state="disabled")
            self.status_var.set("Processing reports...")
            thread = threading.Thread(target=self.run_process)
            thread.start()
    
    def run_process(self):
        """Run the report processing in a separate thread"""
        try:
            main(to_emails=self.selected_emails, status_callback=self.status_callback)
            self.root.after(0, lambda: messagebox.showinfo("Success", "Reports sent successfully!"))
            self.root.after(100, self.root.quit)
        except Exception as e:
            self.root.after(0, lambda: messagebox.showerror("Error", f"An error occurred: {e}"))
            self.root.after(0, lambda: self.send_btn.config(state="normal"))
        finally:
            self.root.after(0, lambda: self.status_var.set("Process complete"))
    
    def status_callback(self, message):
        """Update status from background thread"""
        self.root.after(100, lambda: self.status_var.set(message))
    
    def run(self):
        """Start the UI"""
        self.root.mainloop()
        return self.selected_emails


# Main function
def main(to_emails=None, status_callback=None, send_email=True, upload_to_drive=False):
    """
    Main function to execute the script for multiple reports and send an email with the processed files.
    
    Args:
        to_emails: List of email addresses to send to (optional)
        status_callback: Function to call with status updates (optional)
        send_email: Send processed files via email (optional, default True)
        upload_to_drive: Upload processed files to Google Drive (optional, default False)
    """

    # Clear directory before processing new reports
    clear_save_directory(SAVE_DIRECTORY)

    # List of subject filters for the reports
    subject_filters = [
        "Report: MAL: Chowchilla Womens Prison Abuse - ACTS - AWD - Shield Legal",
        "Report: MAL: Dr Barry Brock SA - AWD - SGGH - Shield Legal",
        "Report: MAL: Dr Derrick Todd Abuse - SGGH - AWD - Shield Legal",
        "Report: MAL: Dr Scott Lee Abuse - SGGH - AWD - Shield Legal",
        "Report: MAL: Mormon Victim Abuse - Dolman - Anapol Weiss - AWD - Shield Legal",
        "Report: MAL: Mormon Victim Abuse - HRSC - AWD - Shield Legal",
        "Report: MAL: NEC - AWD - Wagstaff - Shield Legal",
        "Report: MAL: Paraquat - AWD - Wagstaff - Shield Legal",
        # # 
        "Report: MAL: Paraquat 2 - AWD - Wagstaff - Shield Legal",
        "Report: MAL: San Bernardino County JDC Abuse - SGGH - AWD - Shield Legal",
        "Report: MAL: San Bernardino County JDC Abuse OSOL - SGGH - AWD - Shield Legal",
        "Report: MAL: San Diego County JDC Abuse - SGGH - AWD - Shield Legal",
        "Report: MAL: San Diego County JDC Abuse OSOL - SGGH - AWD - Shield Legal",
        "Report: MAL: Transvaginal Mesh - Anapol Weiss - AWD - Shield Legal",
        "Report: MAL: Video Gaming Sextortion - Anapol Weiss - AWD - 3PL",
        "Report: MAL: Video Gaming Sextortion - Anapol Weiss - AWD - Shield Legal",
        "Report: MAL: Video Gaming Sextortion - Cooper Masterman - AWD - Shield Legal",
        "Report: MAL: Video Gaming Sextortion SEC - Anapol Weiss - AWD - 3PL",
        "Report: MAL: Video Gaming Sextortion SEC - Anapol Weiss - AWD - Anapol Weiss",
        "Report: MAL: Instant Soup Cup Child Burns - BCL - Gomez - Shield Legal",
        ##
        "Report: MAL: Chowchilla Womens Prison Abuse - Oakwood - Oakwood - Shield Legal",
        "Report: MAL: Polinsky Children's Center Abuse 2 - Oakwood - Oakwood - Shield Legal"
        # Add more subject filters here
    ]

    # Folder to save processed files
    attachment_folder = SAVE_DIRECTORY

    # Email details - use provided emails or defaults
    if to_emails is None:
        to_emails = [
            "aidan@tortintakeprofessionals.com",
            "ngaston@tortintakeprofessionals.com",
            "martin@tortintakeprofessionals.com",
            "mclark@tortintakeprofessionals.com",
            "oroman@tortintakeprofessionals.com",
            "pjerome@tortintakeprofessionals.com",
            "esteban@tortintakeprofessionals.com",
            "brittany@tortintakeprofessionals.com",
            "jackson@tortintakeprofessionals.com"
        ]
    email_subject = "Malissa Monday Reports"
    email_body = (
        "Hello,\n\n"
        "Please find attached the processed reports for Cameron & Crump's Monday Reports.\n\n"
        "Best regards,\n"
        "Your Automation Script ʕ•́ᴥ•̀ʔっ♡"
    )

    # Process each report
    for idx, subject_filter in enumerate(subject_filters, 1):
        status_msg = f"Processing report {idx}/{len(subject_filters)}: {subject_filter[:50]}..."
        print(status_msg)
        if status_callback:
            status_callback(status_msg)
        
        file_path = get_report_email(gservice, subject_filter)
        if file_path:
            try:
                # Process the Excel file
                processed_file_path = process_excel(file_path)
                if processed_file_path:
                    rename_sheet(processed_file_path, "Sheet1", "All Fields All Time")
                    add_table_to_sheet(processed_file_path, "All Fields All Time")
                    pivot_sheets = [
                        "Pivot Table"
                        # "Pivot Table Matches Dashboard",
                        # "Pivot Table All Final",
                    ]
                    create_multiple_pivot_tables(processed_file_path, "All Fields All Time", pivot_sheets)
                    format_and_reorder_sheets(
                        processed_file_path,
                        ["All Fields All Time"] + pivot_sheets,
                    )
            except Exception as e:
                print(f"Error processing report: {e}")

        print(f"Finished processing report: {subject_filter}")
        print("-" * 50)

    # Send an email with all processed files
    if send_email:
        try:
            status_msg = "Sending email with attachments..."
            print(status_msg)
            if status_callback:
                status_callback(status_msg)
            send_email_with_attachments(gservice, to_emails, email_subject, email_body, attachment_folder)
            print("Email sent successfully!")
        except Exception as e:
            print(f"Error sending email: {e}")
    else:
        print("Skipping email (send_email=False)")

    # Upload to Google Drive if requested
    if upload_to_drive:
        try:
            upload_msg = "Uploading files to Google Drive..."
            print(upload_msg)
            if status_callback:
                status_callback(upload_msg)
            upload_folder_to_drive(attachment_folder, "Malissa", status_callback)
            print("Files uploaded to Google Drive successfully!")
        except Exception as e:
            error_msg = f"Error uploading to Drive: {e}"
            print(error_msg)
            if status_callback:
                status_callback(error_msg)

    print("All reports processed successfully.")

def launch_ui():
    """Launch the UI for email selection"""
    # Available email addresses
    available_emails = [
        "aidan@tortintakeprofessionals.com",
        "ngaston@tortintakeprofessionals.com",
        "martin@tortintakeprofessionals.com",
        "mclark@tortintakeprofessionals.com",
        "oroman@tortintakeprofessionals.com",
        "pjerome@tortintakeprofessionals.com",
        "esteban@tortintakeprofessionals.com",
        "brittany@tortintakeprofessionals.com",
        "jackson@tortintakeprofessionals.com"
    ]
    
    # Subject filters for reports
    subject_filters = [
        "Report: MAL: Chowchilla Womens Prison Abuse - ACTS - AWD - Shield Legal",
        "Report: MAL: Dr Barry Brock SA - AWD - SGGH - Shield Legal",
        "Report: MAL: Dr Derrick Todd Abuse - SGGH - AWD - Shield Legal",
        # More reports configured in main()
    ]
    
    # Create and run UI
    ui = EmailSenderUI(available_emails, subject_filters)
    ui.run()


if __name__ == "__main__":
    # Launch UI instead of running directly
    launch_ui()


# if you get this error:
# AttributeError: module 'win32com.gen_py.00020813-0000-0000-C000-000000000046x0x1x9' has no attribute 'CLSIDToClassMap'
# goto this folder and delete the folder: C:\Users\Aidan\AppData\Local\Temp\gen_py


