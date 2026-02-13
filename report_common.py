"""
Common functions shared across all Monday report scripts.
Extracted from the 4 individual report scripts to eliminate duplication.
"""

import os
import re
import base64
import shutil
import pandas as pd
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
import win32com.client as win32
import pythoncom

import gmail_auth
from xlsx_fixer import XLSXFixer
from drive_uploader import upload_folder_to_drive
from report_config import REPORTS, get_save_directory, get_default_emails

# Authenticate Gmail at module load (same as original scripts)
gservice = gmail_auth.confirm_auth()


def get_target_date():
    """
    Get today's date if it's Monday, otherwise get the most recent Monday.
    Returns date in YYYY/MM/DD format for Gmail query.
    """
    today = datetime.now().date()
    weekday = today.weekday()  # 0 = Monday, 6 = Sunday

    if weekday == 0:  # Today is Monday
        target_date = today
    else:  # Calculate most recent Monday
        days_since_monday = weekday
        target_date = today - timedelta(days=days_since_monday)

    return target_date.strftime("%Y/%m/%d")


def clear_save_directory(directory):
    """
    Clears all files in the specified directory.
    """
    if os.path.exists(directory):
        shutil.rmtree(directory)
        os.makedirs(directory, exist_ok=True)
        print(f"Cleared directory: {directory}")
    else:
        os.makedirs(directory)
        print(f"Created new directory: {directory}")


def sanitize_filename(filename):
    """Remove characters invalid in Windows filenames."""
    return re.sub(r'[\\/*?:"<>|]', "_", filename)


def get_report_email(gservice, subject_filter, save_directory):
    """Search Gmail for a report email and download its .xlsx attachment."""
    target_date = get_target_date()
    query = f"subject:\"{subject_filter}\" after:{target_date}"
    print(f"Searching for emails after {target_date}")
    results = gservice.users().messages().list(userId="me", q=query).execute()
    messages = results.get("messages", [])

    if not messages:
        print("No emails found with the given subject.")
        return None

    for message in messages:
        msg = gservice.users().messages().get(userId="me", id=message["id"]).execute()
        payload = msg.get("payload", {})
        parts = payload.get("parts", [])

        for part in parts:
            if part.get("filename") and part.get("mimeType") == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
                filename = sanitize_filename(part["filename"])
                body = part.get("body", {})
                if "attachmentId" in body:
                    attachment_id = body["attachmentId"]
                    attachment = gservice.users().messages().attachments().get(
                        userId="me", messageId=message["id"], id=attachment_id
                    ).execute()
                    file_data = base64.urlsafe_b64decode(attachment["data"])

                    # Save the file
                    os.makedirs(save_directory, exist_ok=True)
                    file_path = os.path.join(save_directory, filename)
                    with open(file_path, "wb") as f:
                        f.write(file_data)
                    print(f"File downloaded: {file_path}")
                    return file_path

    print("No attachments found.")
    return None


def process_excel(file_path):
    """
    Fix Excel structure, process data, and format date columns to Excel's 'Short Date'.
    """
    try:
        print("Fixing Excel structure...")
        XLSXFixer.fix_default_col_width(file_path)

        print("Processing Excel data...")
        df = pd.read_excel(file_path)

        # Date columns to format
        date_columns = ["E-Sign Signed Date", "Lead Created Date", "Date of Birth"]
        for col in date_columns:
            if col in df.columns:
                df[col] = pd.to_datetime(df[col], errors="coerce")

        # Save DataFrame back to Excel
        df.to_excel(file_path, index=False, engine="openpyxl")

        # Explicitly apply Excel 'Short Date' formatting using openpyxl
        wb = load_workbook(file_path)
        ws = wb.active

        for col_name in date_columns:
            if col_name in df.columns:
                col_idx = df.columns.get_loc(col_name) + 1  # pandas is 0-based, Excel is 1-based
                for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                    cell = row[0]
                    if cell.value:
                        cell.number_format = 'MM/DD/YYYY'  # Excel Short Date format

        wb.save(file_path)
        print(f"Dates formatted as 'Short Date'. Final file saved: {file_path}")

        return file_path

    except Exception as e:
        print(f"Error processing Excel file: {e}")
        return None


def add_table_to_sheet(file_path, sheet_name):
    """Add a formatted Excel table to the named sheet."""
    try:
        wb = load_workbook(file_path)
        ws = wb[sheet_name]
        max_row, max_col = ws.max_row, ws.max_column
        table_range = f"A1:{ws.cell(row=max_row, column=max_col).coordinate}"
        table = Table(displayName="Table1", ref=table_range)
        style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True, showColumnStripes=True)
        table.tableStyleInfo = style
        ws.add_table(table)
        wb.save(file_path)
        print(f"Table added to sheet: {sheet_name}")

    except Exception as e:
        print(f"Error adding table to sheet: {e}")


def rename_sheet(file_path, old_name, new_name):
    """Rename a sheet in an Excel workbook."""
    try:
        wb = load_workbook(file_path)
        if old_name in wb.sheetnames:
            wb[old_name].title = new_name
            wb.save(file_path)
            print(f"Sheet '{old_name}' renamed to '{new_name}'.")
        else:
            print(f"Sheet '{old_name}' not found in the workbook.")
    except Exception as e:
        print(f"Error renaming sheet: {e}")


def create_multiple_pivot_tables(file_path, data_sheet_name, pivot_sheet_names):
    """Create pivot tables using win32com Excel COM automation."""
    excel = None
    try:
        # Initialize COM for this thread
        pythoncom.CoInitialize()

        excel = win32.gencache.EnsureDispatch("Excel.Application")
        excel.Visible = False
        wb = excel.Workbooks.Open(os.path.abspath(file_path))
        data_sheet = wb.Worksheets(data_sheet_name)

        last_row = data_sheet.UsedRange.Rows.Count
        last_col = data_sheet.UsedRange.Columns.Count

        # Validate that Status column exists (prevents silent pivot failures)
        headers = [data_sheet.Cells(1, col).Value for col in range(1, last_col + 1)]
        if "Status" not in headers:
            raise ValueError("The 'Status' column is missing from the data. Cannot create pivot table.")

        data_range = data_sheet.Range(data_sheet.Cells(1, 1), data_sheet.Cells(last_row, last_col))

        pivot_cache = wb.PivotCaches().Create(SourceType=1, SourceData=data_range)

        for pivot_sheet_name in pivot_sheet_names:
            pivot_sheet = wb.Sheets.Add()
            pivot_sheet.Name = pivot_sheet_name
            pivot_table = pivot_cache.CreatePivotTable(
                TableDestination=pivot_sheet.Cells(1, 1),
                TableName=f"PivotTable_{pivot_sheet_name.replace(' ', '_')}",
            )
            status_field_row = pivot_table.PivotFields("Status")
            status_field_row.Orientation = 1
            status_field_row.Position = 1
            status_field_value = pivot_table.PivotFields("Status")
            status_field_value.Orientation = 4
            status_field_value.Function = -4112  # xlCount
            status_field_value.Name = "Count of Status"

        # Explicitly select only one sheet (deselect grouped sheets)
        wb.Worksheets(data_sheet_name).Activate()
        excel.ActiveWindow.View = 1  # xlNormalView
        excel.ActiveWindow.SelectedSheets(1).Select()

        wb.Save()
        print("Pivot tables created successfully.")

    except Exception as e:
        print(f"Error creating pivot tables: {e}")

    finally:
        if excel is not None:
            excel.Quit()
        # Uninitialize COM
        pythoncom.CoUninitialize()


def format_and_reorder_sheets(file_path, sheet_order):
    """
    Adjust column widths for the first sheet in the order and reorder sheets.
    Ensures only the first sheet is selected upon saving.
    """
    excel = None
    try:
        # Initialize COM for this thread
        pythoncom.CoInitialize()

        excel = win32.gencache.EnsureDispatch("Excel.Application")
        excel.Visible = False
        print(f"Opening workbook for formatting and reordering: {file_path}")

        wb = excel.Workbooks.Open(os.path.abspath(file_path))

        # Verify sheets exist
        existing_sheets = [sheet.Name for sheet in wb.Sheets]
        for sheet in sheet_order:
            if sheet not in existing_sheets:
                raise ValueError(f"Sheet '{sheet}' not found in workbook.")

        # Adjust column widths for the first sheet
        first_sheet_name = sheet_order[0]
        first_sheet = wb.Worksheets(first_sheet_name)
        first_sheet.Activate()

        print(f"Auto-adjusting column widths for '{first_sheet_name}'...")
        first_sheet.UsedRange.Columns.AutoFit()

        # Reorder sheets explicitly
        print("Reordering sheets...")
        for idx, sheet_name in enumerate(sheet_order, start=1):
            wb.Worksheets(sheet_name).Move(Before=wb.Worksheets(idx))

        # Explicitly select only the first sheet and ungroup sheets
        first_sheet.Select()
        excel.ActiveWindow.View = 1  # xlNormalView ensures ungrouped sheets
        excel.ActiveWindow.SelectedSheets(1).Select()

        wb.Save()
        wb.Close()
        print(f"Sheets reordered and column widths adjusted successfully in '{file_path}'.")

    except PermissionError as e:
        print(f"Permission denied error: {e}")
        print("Ensure the Excel file isn't open elsewhere.")
    except Exception as e:
        print(f"Error during formatting and reordering: {e}")
    finally:
        if excel is not None:
            excel.Quit()
        # Uninitialize COM
        pythoncom.CoUninitialize()


def send_email_with_attachments(gservice, to_emails, subject, body, attachment_folder):
    """Send an email with all .xlsx files in a folder as attachments."""
    try:
        if isinstance(to_emails, list):
            to_emails = ", ".join(to_emails)

        message = MIMEMultipart()
        message["to"] = to_emails
        message["subject"] = subject
        message.attach(MIMEText(body, "plain"))

        # Attach all .xlsx files in the directory
        for file_name in os.listdir(attachment_folder):
            if file_name.endswith(".xlsx"):
                file_path = os.path.join(attachment_folder, file_name)
                with open(file_path, "rb") as attachment:
                    part = MIMEBase("application", "octet-stream")
                    part.set_payload(attachment.read())
                encoders.encode_base64(part)
                part.add_header(
                    "Content-Disposition",
                    f"attachment; filename={os.path.basename(file_path)}",
                )
                message.attach(part)

        raw_message = base64.urlsafe_b64encode(message.as_bytes()).decode()
        gservice.users().messages().send(userId="me", body={"raw": raw_message}).execute()
        print("Email sent successfully!")
    except Exception as e:
        print(f"Error sending email: {e}")


def run_report(config_key, to_emails=None, status_callback=None,
               send_email=True, upload_to_drive=False):
    """
    Generic report runner. Replaces the per-script main() function.

    Args:
        config_key: Key into REPORTS dict (e.g., "andy_greg")
        to_emails: Override recipient list (optional)
        status_callback: UI status update callback (optional)
        send_email: Whether to send via email (default True)
        upload_to_drive: Whether to upload to Google Drive (default False)
    """
    config = REPORTS[config_key]
    save_dir = get_save_directory(config_key)

    # Clear directory before processing new reports
    clear_save_directory(save_dir)

    subject_filters = config["subject_filters"]
    pivot_sheets = config["pivot_sheets"]

    if to_emails is None:
        to_emails = get_default_emails(config_key)

    # Process each report
    for idx, subject_filter in enumerate(subject_filters, 1):
        status_msg = f"Processing report {idx}/{len(subject_filters)}: {subject_filter[:50]}..."
        print(status_msg)
        if status_callback:
            status_callback(status_msg)

        file_path = get_report_email(gservice, subject_filter, save_dir)
        if file_path:
            try:
                # Process the Excel file
                processed_file_path = process_excel(file_path)
                if processed_file_path:
                    rename_sheet(processed_file_path, "Sheet1", "All Fields All Time")
                    add_table_to_sheet(processed_file_path, "All Fields All Time")
                    create_multiple_pivot_tables(processed_file_path, "All Fields All Time", pivot_sheets)
                    format_and_reorder_sheets(
                        processed_file_path,
                        ["All Fields All Time"] + pivot_sheets,
                    )
            except Exception as e:
                print(f"Error processing report: {e}")

        print(f"Finished processing report: {subject_filter}")
        print("-" * 50)

    # Send email if enabled
    if send_email:
        try:
            status_msg = "Sending email with attachments..."
            print(status_msg)
            if status_callback:
                status_callback(status_msg)
            send_email_with_attachments(
                gservice, to_emails,
                config["email_subject"], config["email_body"], save_dir
            )
            print("Email sent successfully!")
        except Exception as e:
            print(f"Error sending email: {e}")
    else:
        print("Skipping email (send_email=False)")

    # Upload to Google Drive if requested
    if upload_to_drive:
        try:
            upload_msg = "Uploading files to Google Drive..."
            print(upload_msg)
            if status_callback:
                status_callback(upload_msg)
            upload_folder_to_drive(save_dir, config["drive_folder_name"], status_callback)
            print("Files uploaded to Google Drive successfully!")
        except Exception as e:
            error_msg = f"Error uploading to Drive: {e}"
            print(error_msg)
            if status_callback:
                status_callback(error_msg)

    print("All reports processed successfully.")
